from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = Path(r"C:\Users\maksy\Downloads\keyword_gap_biomasaportal (5).xlsx")
OUTPUT_PATH = ROOT_DIR / "data" / "editorial" / "articles-seed.json"


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    return normalized.strip("-")


def clean_text(value: object) -> str:
    return str(value or "").strip()


def main() -> None:
    workbook = load_workbook(DEFAULT_INPUT, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    records: list[dict[str, object]] = []

    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        keyword, title, meta_title, meta_description, html_article, faq_schema, image_prompts = row[:7]

        if not keyword and not title:
            continue

        normalized_title = clean_text(title)
        slug = slugify(normalized_title or clean_text(keyword) or f"artykul-{index}")

        records.append(
            {
                "order": index,
                "sourceRow": index + 1,
                "keyword": clean_text(keyword),
                "title": normalized_title,
                "slug": slug,
                "path": f"/{slug}/",
                "metaTitle": clean_text(meta_title),
                "metaDescription": clean_text(meta_description),
                "htmlArticle": clean_text(html_article),
                "faqSchema": clean_text(faq_schema),
                "imagePrompts": clean_text(image_prompts),
                "initialStatus": "already_published" if index <= 10 else "queued",
            }
        )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "sourceWorkbook": str(DEFAULT_INPUT),
                "sheetName": worksheet.title,
                "generatedCount": len(records),
                "items": records,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Saved {len(records)} editorial items to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
